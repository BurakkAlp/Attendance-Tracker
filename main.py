import tkinter as tk #for using tkinter
from tkinter import ttk, filedialog #ttk for advanced tkinter widgets and filedialog for directory selections
import openpyxl #for handling excel

class Student: # student class is created, each student object has given id, name, dept and section
    def __init__(self, ID, name, dept, section):
        student_count = 0
        self.name = name
        self.ID = ID
        self.dept = dept
        self.section = section
        student_count += 1 # this is used for to count how much object has been created

    def print_student(self): # for printing students
        print(f"ID: {self.ID}, Name: {self.name}, Dept: {self.dept}, Section: {self.section}")

    def getName(self): #get function used to retrieve the values of private attributes
        return self.name

    def getID(self):
        return self.ID

    def getDept(self):
        return self.dept

    def getSection(self):
        return self.section

class StudentList: #studentlis class is created with only empty list
    def __init__(self):
        self.student_list = []

    def add_student(self, student): #for appending students
        self.student_list.append(student)

    def remove_student(self, student): #for removing students
        self.student_list.remove(student)

    def print_student_list(self): #for printing all students
        for student in self.student_list:
            student.print_student()

    def read_students_from_excel(self, file_path): #this method is for taking data inside the excel file
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            student_list = StudentList()

            for row in sheet.iter_rows(values_only=True, min_row=2):
                if all(row):
                    if len(row) >= 4:
                        student = Student(row[0], row[1], row[2], row[3])
                        student_list.add_student(student)
                    else:
                        print("Warning: Incomplete data in row, skipping:", row)
                else:
                    print("Warning: Empty row, skipping.")

            workbook.close()
            return student_list
        except Exception as e:
            print("Error reading Excel file:", e)
            return None

    def write_students_to_excel(self, section, week_number, file_type=".xlsx"): #this method is for writing data to excel file
        filename = f"{section}, Week{week_number}{file_type}"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Name", "Dep"])

        for student in self.student_list:
            sheet.append([student.getID(), student.getName(), student.getDept()])

        workbook.save(filename)


class Gui(tk.Tk): #every single widget has been created in this class and some methods for giving buttons functionality
    def __init__(self):
        super().__init__()
        self.title("tk") #title
        self.geometry("645x315+400+350")#lenght and width of window
        self.student_list = StudentList()

        #labels
        self.label1 = tk.Label(text="AttendanceKeeper v1.0", font="Times 20")
        self.label1.grid(row=0, column=0, columnspan=3, padx=180)

        self.label2 = tk.Label(text="Select student list excel file:", font="Times 12")
        self.label2.grid(row=1, column=0, sticky="w", padx=(30, 0), pady=10)

        self.label3 = tk.Label(text="     Select a student:", font="Times 12")
        self.label3.grid(row=2, column=0, sticky="w", padx=(30, 0), pady=10)

        self.label4 = tk.Label(text="      Section:", font="Times 12")
        self.label4.grid(row=2, column=1, sticky="w", padx=(0, 30), pady=10)

        self.label5 = tk.Label(text="       Attended Students:", font="Times 12")
        self.label5.grid(row=2, column=2, sticky="w", pady=10)

        self.label6=tk.Label(text="Please select file type:")
        self.label6.grid(row=6, column=0, sticky="w", pady=10)

        self.label7=tk.Label(text="Please enter week:")
        self.label7.grid(row=6, column=1, sticky="we", pady=10)

        #entry

        self.entry=tk.Entry(width=15)
        self.entry.grid(row=6,column=2,sticky="w")
        #datas inside combobox
        lectures = ["AP 01", "AP 02", "AP 03", "AP 04", "AP 05", "AP 06", "AP 07", "AP 08", "AP 09", "AP 10",
                    "AP 11", "AP 12", "AP 13", "AP 14", "AP 15", "AP 16", "AP 17", "AP 18", "AP 19", "AP 20"]
        file_type = [".xls", ".csv", ".txt"]

        #comboboxes
        self.combobox1 = ttk.Combobox(values=lectures, height=2, width= 15)
        self.combobox1.set("AP 01")
        self.combobox1.grid(row=3, column=1, padx=25, pady=1, sticky="nw")

        self.combobox2 = ttk.Combobox(values=file_type, height=2, width=5)
        self.combobox2.set(".txt")
        self.combobox2.grid(row=6, column=0, sticky="e", pady=10)

        #listboxes
        self.listbox1 = tk.Listbox(selectmode=tk.MULTIPLE, height=8)
        self.listbox1.grid(row=3, column=0, rowspan=3, sticky="nsew", padx=(10, 0))

        self.listbox2 = tk.Listbox(selectmode=tk.MULTIPLE, height=8, width=30)
        self.listbox2.grid(row=3, column=2, rowspan=3, sticky="ew")

        def write_to_listbox1_from_excel(): #the data read from Excel is added to listbox1 and initially the AP 01 section is displayed
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            selected_section = self.combobox1.get()
            if file_path:
                self.listbox1.delete(0, tk.END)
                self.listbox2.delete(0, tk.END)
                student_list = StudentList().read_students_from_excel(file_path)
                if student_list:
                    for student in student_list.student_list:
                        if student.getSection() == selected_section:
                            name = student.getName()
                            ID = student.getID()
                            self.listbox1.insert(tk.END, f"{name}, {ID}")

        def add_selected_students_to_listbox2(): #add button functionality
            selected_indices = self.listbox1.curselection()
            for index in selected_indices[::-1]:
                student = self.listbox1.get(index)
                self.listbox1.delete(index)
                self.listbox2.insert(tk.END, student)

        def remove_selected_students_from_listbox2(): #remove button functionality
            selected_indices = self.listbox2.curselection()
            for index in selected_indices[::-1]:
                student = self.listbox2.get(index)
                self.listbox2.delete(index)
                self.listbox1.insert(tk.END, student)

        def export_students_to_txt(): #when the file type is selected txt and user presses export as file this method is creates text file
            selected_section = self.combobox1.get()
            week_number = self.entry.get()
            file_type = self.combobox2.get()
            filename = f"{selected_section}, Week{week_number} {file_type}"
            with open(filename, "w",encoding="utf-8") as file:
                for student in self.student_list.student_list:
                    file.write(f"{student.getID()}, {student.getName()}, {student.getDept()}\n")

        def export_students_to_excel(): #when the file type is selected xlsx and user presses export as file this method is creates excel file
            selected_section = self.combobox1.get()
            week_number = self.entry.get()
            file_type = self.combobox2.get()
            students = [self.listbox2.get(idx) for idx in range(self.listbox2.size())]

            for student_str in students:
                student_info = student_str.split(", ")
                student = Student(student_info[1], student_info[0], "", selected_section)  # Boş departman bilgisi
                self.student_list.add_student(student)
            try:
                if file_type == ".txt":
                    export_students_to_txt()
                else:
                    self.student_list.write_students_to_excel(selected_section, week_number, file_type)
                self.listbox2.delete(0, tk.END)
                self.student_list.student_list.clear()
            except Exception as e:
                print("Error exporting data:", e)

        #buttons
        self.buton1 = tk.Button(text="  Import List", font="Times 12", command=write_to_listbox1_from_excel)
        self.buton1.grid(row=1, column=1, padx=(50, 60), pady=10, sticky="nswe")

        self.buton2 = tk.Button(text="  Add =>", font="Times 12", command=add_selected_students_to_listbox2)
        self.buton2.grid(row=4, column=1, padx=25, pady=10, sticky="nswe")

        self.buton3 = tk.Button(text="  <= Remove", font="Times 12", command=remove_selected_students_from_listbox2)
        self.buton3.grid(row=5, column=1, padx=25, pady=10, sticky="nswe")

        self.buton4 = tk.Button(text="Export as file", font="Times 12", command=export_students_to_excel)
        self.buton4.grid(row=6, column=2, padx=(0, 12), pady=10, sticky="e")

def main(): #main function
    app = Gui()
    app.mainloop()
main()

